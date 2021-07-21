from selenium import webdriver
from selenium.webdriver.support.select import Select
from  selenium.webdriver import ActionChains
import openpyxl
class Base:
    def launch_browser(self):
        self.driver = webdriver.Chrome(executable_path=r"C:\Users\DELL\PycharmProjects\SELENIUM\Driver\chromedriver.exe")
        return self.driver

    #WebDriver
    def window_maximize(self):
        self.driver.maximize_window()

    def window_minimize(self):
        self.driver.minimize_window()

    def load_url(self,url):
        self.driver.get(url)

    def browser_quit(self):
        self.driver.quit()

    def close_browser(self):
        self.driver.close()

    def capture_screen_shot(self,name):
        self.driver.save_screenshot(name)

    def browser_refresh(self):
        self.driver.refresh()

    def move_forward(self):
        self.driver.forward()

    def move_back(self):
        self.driver.back()

    def page_url(self):
        url = self.driver.current_url
        return url

    def get_current_window_id(self):
        parent_id = self.driver.current_window_handle
        return parent_id

    def get_all_windows_id(self):
        all_window_id = self.driver.window_handles
        return all_window_id

    def page_title(self):
        title = self.driver.title
        return title

    #WebElement
    def send_txt(self,element,data):
        element.send_keys(data)

    def btn_click(self,e):
        e.click()

    def attribute_value(self,element,value):
        attribute = element.get_attribute(value)
        return attribute

    def displayed(self,element):
        d = element.is_displayed()
        return d

    def selected(self,element):
        e = element.is_selected()
        return  e

    def enabled(self,element):
        f = element.is_enabled()
        return f

    def print_text(self,element):
        t1 = element.text
        return t1
    def rtn_text(self,ele):
        for i in ele:
            h = i.text
            print(h)


    #alert
    def switch_alert(self):
        al = self.driver.switch_to.alert
        return al

    def alert_ok(self):
        al = self.switch_alert()
        al.accept()

    def alert_cancel(self):
        al = self.switch_alert()
        al.dismiss()

    def fill_alert(self,data):
        al = self.switch_alert()
        al.send_keys(data)

    def get_text_from_alert(self):
        al = self.switch_alert()
        t = al.text
        return t

    #Actionchains
    def mouse_over(self,element):
        a = ActionChains(self.driver)
        a.move_to_element(element).perform()

    def drag_drop(self,src,des):
        a = ActionChains(self.driver)
        a.drag_and_drop(src,des).perform()

    def right_click(self,ele):
        a = ActionChains(self.driver)
        a.context_click(ele).perform()

    def double_tab(self,e):
        a = ActionChains(self.driver)
        a.double_click(e).perform()

    #switch to frame
    def switch_frame_by_id(self,id):
        self.driver.switch_to.frame(id)

    def switch_frame_by_name(self,name):
        self.driver.switch_to.frame(name)

    def switch_frame_by_index(self,index):
        self.driver.switch_to.frame(index)

    def switch_frame_by_element(self,element):
        self.driver.switch_to.frame(element)

    def switch_parent_frame(self):
        self.driver.switch_to.parent_frame()

    #select

    def dd_by_index(self,e,index):
        s = Select(e)
        s.select_by_index(index)

    def dd_by_value(self,ele,value):
        s = Select(ele)
        s.select_by_value(value)

    def dd_by_visible_txt(self,el,txt):
        s = Select(el)
        s.select_by_visible_text(txt)

    def deselect_index(self,e,index):
        s = Select(e)
        s.deselect_by_index(index)

    def deselect_value(self,el,value):
        s = Select(el)
        s.deselect_by_value(value)

    def deselect_visibletxt(self,ele,txt):
        s = Select(ele)
        s.deselect_by_visible_text(txt)

    def to_deselect_all_in_dd(self,element):
        s = Select(element)
        s.deselect_all()

    def dd_option(self,element):
        s = Select(element)
        op = s.options
        return op



    def dd_fist_option(self,e):
        s = Select(e)
        f = s.first_selected_option
        return f

    def dd_all_select_option(self,ele):
        s = Select(ele)
        o = s.all_selected_options
        return o

    def dd_multiple(self,element):
        s = Select(element)
        p = s.is_multiple
        return p
    def get_data_from_excel(self,loc,row,cell):
        excel_loc = r"C:\Users\DELL\PycharmProjects\selenium\Excel\adactin.xlsx"
        w = openpyxl.load_workbook(excel_loc)
        sheet = w.active
        c = sheet.cell(row,cell)
        v = c.value
        return v
    def store_data_in_excel(self,ele,cell_no):
        excel_loc = "IPHONELIST.xlsx"
        w = openpyxl.Workbook()
        sheet = w.create_sheet("iphone", 0)
        for i in ele:
            row = sheet.max_row
            s = sheet.cell(row + 1, cell_no)
            each_element = i.text
            s.value = each_element
            w.save(excel_loc)










